from typing import Tuple
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from .models import Order, STATUSES


def upsert_order_from_row(db: Session, row: dict) -> Tuple[bool, bool]:
    """Return (created, updated) flags."""
    order_no = str(row.get("order_no") or "").strip()
    if not order_no:
        return (False, False)

    group_code = (row.get("group_code") or None) or None
    status = row.get("status") or STATUSES[0]
    if status not in STATUSES:
        status = STATUSES[0]
    weight = row.get("weight_kg")
    try:
        weight = float(weight) if weight is not None and weight != "" else None
    except Exception:
        weight = None
    fee = row.get("shipping_fee")
    try:
        fee = float(fee) if fee is not None and fee != "" else None
    except Exception:
        fee = None

    created = False
    updated = False
    obj = db.query(Order).filter(Order.order_no == order_no).one_or_none()
    if obj is None:
        obj = Order(order_no=order_no)
        created = True

    # set values
    obj.group_code = group_code
    obj.status = status
    if weight is not None:
        obj.weight_kg = weight
    if fee is not None:
        obj.shipping_fee = fee

    db.add(obj)
    db.flush()
    if not created:
        updated = True
    return (created, updated)


def import_excel(db: Session, file_path: str) -> dict:
    wb = load_workbook(filename=file_path)
    ws = wb.active
    # Expect headers in first row: order_no, group_code, weight_kg, status, shipping_fee
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))[0:5]]
    # Fallback mapping
    expected = ["order_no", "group_code", "weight_kg", "status", "shipping_fee"]
    if any(h not in expected for h in headers):
        headers = expected

    created = 0
    updated = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        data = {headers[i]: r[i] if i < len(r) else None for i in range(len(headers))}
        c, u = upsert_order_from_row(db, data)
        created += (1 if c else 0)
        updated += (1 if u and not c else 0)

    db.commit()
    return {"created": created, "updated": updated}

